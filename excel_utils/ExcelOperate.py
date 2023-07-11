from openpyxl.utils import column_index_from_string as Col2Int
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter
import openpyxl
import numbers
import os
import re
# 操作excel
def generate_get_set_methods(*variable_names):
    for name in variable_names:
        def getter(self):
            return getattr(self, '_' + name)

        def setter(self, value):
            setattr(self, '_' + name, value)

        yield property(getter, setter)

class ExcelOperate:
    def __init__(self, data: dict = None)  -> None:

        self.worksheet = None
        self.filename = os.path.abspath(data.get('filename', ''))
        self.sheet = data.get('sheet', '')
        self.data_range = data.get('data_range', '')
        self.data_pair = data.get('data_pair', [])

        if not all([self.filename, self.sheet, self.data_range, self.data_pair]):
            raise ValueError("Invalid input data")
        # excel所有信息
        wb_id1 = openpyxl.load_workbook(self.filename, data_only=True)
        if isinstance(self.sheet, numbers.Number):
            sheet_list = wb_id1.sheetnames
            self.sheet_info = wb_id1[sheet_list[int(self.sheet)-1]]
        else:
            self.sheet_info = wb_id1[self.sheet]
        wb_id1 = None
        # 获取excel所有范围
        self._range_min, self._range_max = self.get_range_value()

        # 读取所有的数据,和全部转为字符串的数据
        self._all_data_list, self._all_data_string_list = self.read_excel_data()

        # 读取需要判别的数据
        self._data_list, self._data_dict = self.read_simple_data()

        # 读取需要判别的字符串数据
        self._data_string_list, self._data_string_dict = self.read_simple_data()

        # 清空内存
        self.sheet_info = ''

    def get_range_value(self):
        '''获取要读取的范围值
        Returns: tuple

        '''

        # 将有效范围转换为 (min_row, min_col, max_row, max_col) 的元组
        range_min, range_max = self.sheet_info.calculate_dimension().split(':')
        range_min_col, range_min_row = coordinate_from_string(range_min)
        range_max_col, range_max_row = coordinate_from_string(range_max)
        if self.data_range == '*':
            pass
        else:
            # 获取单元格最小值
            range_min, range_max = self.data_range.split(":")
            if range_min in ['*']:
                pass
            else:
                range_min_col, range_min_row = range_min.split(",")
                if range_min_col in ['min', '*']:
                    range_min_col = get_column_letter(self.sheet_info.min_column)
                if range_min_row in ['min', '*']:
                    range_min_row = Col2Int(get_column_letter(self.sheet_info.min_row))

            # 获取单元格最大值
            if range_max in ['*']:
                pass
            else:
                range_max_col, range_max_row = range_max.split(",")
                if range_max_col in ['max','*']:
                    range_max_col = get_column_letter(self.sheet_info.max_column)
                if range_max_row in ['max','*']:
                    range_max_row = Col2Int(get_column_letter(self.sheet_info.max_row))

        range_min = str(range_min_col) + str(range_min_row)
        range_max = str(range_max_col) + str(range_max_row)
        return range_min, range_max

    '''读取范围内的所有数据'''
    def read_excel_data(self, *args, **kwargs):
        # 打开文件
        wb_id1 = openpyxl.load_workbook(self.filename, data_only=True)
        # 打开sheet
        if isinstance(self.sheet, numbers.Number):
            sheet_list = wb_id1.sheetnames
            sheetDataSet1 = wb_id1[sheet_list[int(self.sheet)-1]]
        else:
            sheetDataSet1 = wb_id1[self.sheet]
        data_list = []
        data_string_list = []
        for i in sheetDataSet1[self._range_min:self._range_max]:
            tmp_list = []
            tmp_string_list = []
            for j in i:
                tmp_list.append(j.value)
                tmp_string_list.append(str(j.value).replace(',','，'))
            data_list.append(tmp_list)
            data_string_list.append(tmp_string_list)
        return data_list, data_string_list

    '''读取范围内的所有数据，全部转为字符串'''
    def read_excel_data_for_string(self, *args, **kwargs):
        # 打开文件
        wb_id1 = openpyxl.load_workbook(self.filename, data_only=True)
        # 打开sheet
        sheetDataSet1 = wb_id1[self.sheet]
        data_list = []
        for i in sheetDataSet1[self._range_min:self._range_max]:
            tmp_list = []
            for j in i:
                tmp_list.append(str(j.value))
            data_list.append(tmp_list)
        return data_list

    '''读取用到的数据'''
    def read_simple_data(self):
        one_col, one_row = coordinate_from_string(self._range_min)
        data_list:list = []
        data_dict:list = []
        # Excel 列的正则表达式
        pattern = r"^[A-Z]{1,3}$"
        if isinstance(self.data_pair, dict):
            for row_data in self._all_data_list:
                row_list:list = []
                row_dict:dict = {}
                for data_pair_col_key, data_pair_col_value in self.data_pair.items():
                    if re.match(pattern, str(data_pair_col_value)):
                        row_list.append(row_data[Col2Int(data_pair_col_value) - Col2Int(one_col)])
                        row_dict[str(data_pair_col_key)] = row_data[Col2Int(data_pair_col_value) - Col2Int(one_col)]
                    else:
                        row_dict[str(data_pair_col_key)] = data_pair_col_value
                data_list.append(row_list)
                data_dict.append(row_dict)
        elif isinstance(self.data_pair, list):
            for row_data in self._all_data_list:
                tmp_list = [row_data[Col2Int(data_pair_col) - Col2Int(one_col)] for data_pair_col in self.data_pair]
                tmp_dict = {str(data_pair_col):row_data[Col2Int(data_pair_col) - Col2Int(one_col)] for data_pair_col in self.data_pair}
                data_list.append(tmp_list)
                data_dict.append(tmp_dict)
        return data_list, data_dict

    '''返回列表中有同样值的数据'''
    def repeat_value(self) -> dict:
        dict_char_tmp = {i[0]: self._data_list.count(i) for i in self._data_list if self._data_list.count(i) > 1}
        return dict_char_tmp

    '''List数据插入excel'''
    @classmethod
    def write_list_excel(cls, result_list, file_name='output', header=None):
        if header:
            header = header
        else:
            if len(result_list) > 1:
                header = [get_column_letter(i+1) for i in range(len(result_list[0]))]
            else:
                header = [get_column_letter(i+1) for i in range(len(result_list))]

        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='Sheet1',index=0)
        for h_col in range(1, len(header)+1):
            _ = sheet.cell(column=h_col, row=1, value="{0}".format(header[h_col-1]))
        i=1
        for result in result_list:
            j = 1
            for v in result:
                _ = sheet.cell(row=i,column=j,value="{}".format(v))
                j+=1
            i+=1
        wb.save(filename="{}".format(file_name+'.xlsx'))
        print("写入数据成功{}".format(os.getcwd()+'\\'+file_name+'.xlsx'))

    '''Dict数据插入excel'''
    @classmethod
    def write_dict_excel(cls, result_dict:dict, file_name='output', header=None):
        if header:
            header = header
        else:
            if len(list(result_dict.values())[0]) > 1:
                header = [get_column_letter(i+1) for i in range(len(list(result_dict.values())[0])+1)]
            else:
                header = [get_column_letter(i+1) for i in range(len(list(result_dict.values())[0])+1)]
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='Sheet1', index=0)
        # for h_col in range(1, len(header) + 1):
        #     _ = sheet.cell(column=h_col, row=1, value="{0}".format(header[h_col - 1]))
        i = 1
        for k,v in result_dict.items():
            _ = sheet.cell(row=i, column=1, value="{}".format(k))
            j = 2
            for value in v:
                _ = sheet.cell(row=i, column=j, value="{}".format(value))
                j += 1
            i += 1
        wb.save(filename="{}".format(file_name + '.xlsx'))
        print("写入数据成功{}".format(os.getcwd()+'\\'+file_name+'.xlsx'))

    '''list<dict>数据插入excel'''
    @classmethod
    def write_list_dict_excel(cls, result_list, file_name='output', header=None,sheetName='Sheet1'):
        if header:
            header = header
        else:
            if len(result_list) > 1:
                header = [i for i in result_list[0].keys()]
            # else:
            #     header = [get_column_letter(i + 1) for i in range(len(result_list))]

        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title=sheetName, index=0)
        for h_col in range(1, len(header) + 1):
            _ = sheet.cell(column=h_col, row=1, value="{0}".format(header[h_col - 1]))
        i = 2
        for result in result_list:
            j = 1
            for k,v in result.items():
                _ = sheet.cell(row=i, column=j, value="{}".format(v))
                j += 1
            i += 1
        wb.save(filename="{}".format(file_name + '.xlsx'))
        print("写入数据成功{}".format(os.getcwd() + '\\' + file_name + '.xlsx'))

    def process_all_content(self, pattern:[str,re], repl: [str,re]) -> None:
        """提取出来的所有内容替换操作（根据正则匹配模式，替换为某个内容）

        Args：
            pattern: 需要匹配的正则表达式或者内容
            repl: 替换的正则表达式或者目标内容

        Returns:
            None
        """
        for item in self._data_list:
            for i,item_content in enumerate(item):
                item[i] = re.sub(pattern, repl, item_content)

    def iterate_list(self, items):
        """遍历列表中的每一项内容

        Args:
            items: 需要遍历的内容

        Returns:
            返回一个生成器，可以遍历每一项内容
        """
        for item in items:
            if isinstance(item, (list, tuple, set)):
                yield from self.iterate_list(item)
            else:
                yield item

    def get_data_list(self):
        output = ""
        for content in self.iterate_list(self._data_list):
            output += content + "\n"
        return output

    def get_data_string_list(self):
        output = ""
        for content in self.iterate_list(self._data_string_list):
            output += content + "\n"
        return output

    def __repr__(self):
        output = ""
        for content in self.iterate_list(self._data_list):
            output += content + "\n"
        return output

    @property
    def range_min(self):
        return self._range_min

    @property
    def range_max(self):
        return self._range_max

    @property
    def all_data_list(self):
        return self._all_data_list

    @property
    def all_data_string_list(self):
        return self._all_data_string_list

    @property
    def data_list(self):
        return self._data_list

    @property
    def data_dict(self):
        return self._data_dict

    @property
    def data_string_list(self):
        return self._data_string_list

    @property
    def data_string_dict(self):
        return self._data_string_dict